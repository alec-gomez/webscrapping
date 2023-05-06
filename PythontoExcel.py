import openpyxl as xl 
from openpyxl.styles import Font 

wb= xl.Workbook() 

ws= wb.active 

ws.title= 'First Sheet' 

wb.create_sheet(index=1, title='Second Sheet') 

ws['A1'] = 'Invoice' 

ws['A1'].font= Font(name='Times New Roman', size=24, bold=True, italic=False) 

myfont= Font(name='Times New Roman', size=24, bold=True, italic=False) 

ws['A1'].font= myfont 

ws['A2']= 'Tires' 
ws['A3']= 'Brakes' 
ws['A4']= 'Alignment' 

ws.merge_cells('A1:B1') 

ws['B2']= 450 
ws['B3']= 225 
ws['B4']= 150 

ws['A8']= 'Total' 

ws['A8'].font= myfont

ws['B8']= '=SUM(B2:B4)' 

#ws.column_dimensions['A']= 25
#read the exel file- 'Producereport.xlsx' that you created earlier 
#write all the contents of this file to 'Second Sheet' in the current workbook 

#display the grand total and average 'amt sold' and 'total' 
#at the bottom of the list along with appropriate labels 

write_sheet= wb['Second Sheet'] 

read_wb=xl.load_workbook('ProduceReport.xlsx') 
read_ws= read_wb['ProduceReport'] 


maxC=read_ws.max_column
maxR= read_ws.max_row

#just puts headers in the excel file 
write_sheet['A1']= 'Produce' 
write_sheet['B1']= 'Cost Per Pound' 
write_sheet['C1']= 'Amount Sold' 
write_sheet['D1']= 'Total' 

#help to know what column you are working with 
write_row= 2
write_colA= 1 
write_colB= 2
write_colC= 3
write_colD= 4 

for currentrow in read_ws.iter_rows(min_row=2, max_row= maxR, max_col= maxC): 
    name= currentrow[0].value
    cost= float(currentrow[1].value) 
    amt_sold= float(currentrow[2].value) 
    total= float(currentrow[3].value) 

#to write to sheet just needs a row and column number (write_row increments to go to next row) 
    write_sheet.cell(write_row, write_colA).value= name   
    write_sheet.cell(write_row, write_colB).value= cost 
    write_sheet.cell(write_row, write_colC).value= amt_sold 
    write_sheet.cell(write_row, write_colD).value= total  

    write_row += 1 

#summery info (done with for loop) 
summary_row= write_row + 1 

write_sheet['B' + str(summary_row)] = 'Total' 
#change font for above total summary ^ 
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True) 
#calcualtion for total 
write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(write_row) + ')' 
                                        #^ same as SUM(C2:C42) 
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(write_row) + ')'                                    
                                        #^ same as^SUM(D2:D42) 

summary_row += 1 

write_sheet['B' + str(summary_row)] = 'Average' 
write_sheet['B' + str(summary_row)].font= Font(size=16, bold=True) 

write_sheet['C' + str(summary_row)] = '=Average(C2:C' + str(write_row) + ')' 
write_sheet['D' + str(summary_row)] = '=Average(D2:D' + str(write_row) + ')' 

write_sheet.column_dimensions['A'].width= 16 
write_sheet.column_dimensions['B'].width= 15 
write_sheet.column_dimensions['C'].width= 15 
write_sheet.column_dimensions['D'].width= 15

#formatting the numbers 
for cell in write_sheet["C:C"]: 
    cell.number_format= '#,##0' 

for cell in write_sheet["D:D"]: 
    cell.number_format= u'"$ "#,##0.00' 




#save workbook 
wb.save('PythontoExcel.xlsx') 
